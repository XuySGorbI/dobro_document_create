from tkinter import filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from myparser import pars_dob  # Импорт функции `pars_dob` из модуля `myparser`.
import re
from datetime import datetime

# Класс для парсинга данных и работы с Excel
class dobro_parser:
    """
    Класс для парсинга данных и работы с Excel.
    """

    # Путь к файлу Excel (инициализируется как None)
    _file_path = None
    
    # Словарь для преобразования русских месяцев в английские (используется для обработки дат)
    months = {
        'января': 'January',
        'февраля': 'February',
        'марта': 'March',
        'апреля': 'April',
        'мая': 'May',
        'июня': 'June',
        'июля': 'July',
        'августа': 'August',
        'сентября': 'September',
        'октября': 'October',
        'ноября': 'November',
        'декабря': 'December'
    }

    # Функция для извлечения данных из словаря и подготовки строки для записи в Excel
    def extract_data(self, data):
        """
        Извлекает и преобразует данные из словаря, полученного парсером страницы.
        Формирует структуру для дальнейшей записи в Excel.
        :param data: Словарь, полученный из функции pars_dob.
        :return: Словарь с преобразованными данными.
        """
        event_title = data['EventInfo_event-title__k6Fsy d-none d-md-block']  # Название события
        project = data['EventInfo_event-partner__mHSXd d-none d-md-block']  # Партнёр события
        location = data['CardTypes_card-location__title__uqLH2']  # Место проведения
        time_info = data['CardTypes_card-time__title__b3zsJ']  # Дата и время
        vacancies = data['EventVacanciesTab_tab__ePxnH']  # Вакансии
        url = data['url']  # URL события

        # Разделение строки времени на дату и время
        date_part, time_part = time_info.split(',')
        day, month_rus, year = date_part.strip().split()  # Извлекаем день, месяц и год
        month = self.months[month_rus]  # Преобразуем название месяца в английское
        date = datetime.strptime(f"{day} {month} {year}", '%d %B %Y').strftime('%d.%m.%Y')  # Форматируем дату
        time_range = time_part.strip()  # Убираем лишние пробелы вокруг времени

        # Обработка вакансий (подсчёт волонтёров и благополучателей)
        volunteers = 0
        beneficiaries = 0
        for vacancy in vacancies.split('$'):  # Каждая вакансия разделена символом `$`
            try:
                matches = re.findall(r'(\d+)из\d+', vacancy)  # Находим числа в формате `N из M`
                for match in matches:
                    number = int(match)
                    if 'участник' not in vacancy:  # Если слово "участник" отсутствует, считаем волонтёров
                        volunteers += number
                    else:  # Иначе считаем благополучателей
                        beneficiaries += number
            except ValueError:
                continue

        # Извлечение названия проекта (после символа `#`)
        project_name = project.split('#')[-1]

        return {
            'date': date,
            'time_range': time_range,
            'event_title': event_title,
            'project_name': project_name,
            'location': location,
            'volunteers': volunteers,
            'beneficiaries': beneficiaries,
            'url': url
        }

    # Создание новой строки в Excel
    def create_excel_row(self, data):
        """
        Добавляет новую строку с данными о мероприятии в Excel-файл.
        Если файл не существует — создаёт новый и добавляет заголовки.
        Вставляет формулы для вычисления полугодия, квартала, месяца, часов и человеко-часов.
        :param data: Словарь с данными для записи (результат extract_data).
        """
        extracted_data = self.extract_data(data)
        date = extracted_data['date']
        time_range = extracted_data['time_range']
        event_title = extracted_data['event_title']
        project_name = extracted_data['project_name']
        location = extracted_data['location']
        volunteers = extracted_data['volunteers']
        beneficiaries = extracted_data['beneficiaries']
        url = extracted_data['url']

        # Попытка открыть существующую книгу или создать новую
        try:
            wb = load_workbook(self._file_path)  # Загружаем существующую книгу
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()  # Создаём новую книгу
            ws = wb.active
            # Добавляем заголовки
            headers = ['Полугодие', 'квартал', 'месяц', 'дата', 'часы мероприятия', 'мероприятие', 
                       'Проект', 'количество волонтеров', 'количество благополучателей', 'место проведения', 
                       'краткое описание', 'Ссылка', 'время проведения', 'общее количество часов волонтёров']
            ws.append(headers)

        # Определяем следующую доступную строку
        last_row = ws.max_row + 1

        # Запись данных в ячейки
        ws[f'D{last_row}'] = date
        ws[f'E{last_row}'] = f'=ТЕКСТ((ВРЕМЗНАЧ(ПСТР(M{last_row}, НАЙТИ("-", M{last_row}) + 2, 5)) - ВРЕМЗНАЧ(ЛЕВСИМВ(M{last_row}, НАЙТИ("-", M{last_row}) - 2))), "ч")'
        ws[f'F{last_row}'] = event_title
        ws[f'G{last_row}'] = project_name
        ws[f'H{last_row}'] = volunteers
        ws[f'I{last_row}'] = beneficiaries
        ws[f'J{last_row}'] = location
        ws[f'L{last_row}'] = url
        ws[f'M{last_row}'] = time_range
        ws[f'N{last_row}'] = f'=H{last_row} * E{last_row}'

        # Формулы для полугодия, квартала и месяца
        ws[f'A{last_row}'] = f'=ЕСЛИ(МЕСЯЦ(D{last_row}) <= 6, 1, 2)'
        ws[f'B{last_row}'] = f'=ОКРУГЛВВЕРХ(МЕСЯЦ(D{last_row})/3, 0)'
        ws[f'C{last_row}'] = f'=ПРОПИСН(ТЕКСТ(D{last_row}, "ММММ"))'

        # Сохранение изменений
        wb.save(self._file_path)

    # Обработка нажатия кнопки
    def for_button_pars(self, url_entry, error_label, tree):
        """
        Обрабатывает нажатие кнопки "добавить" или автоматическую вставку по ссылке.
        Получает ссылку, парсит данные, добавляет строку в Excel и обновляет таблицу в интерфейсе.
        :param url_entry: Поле ввода или строка с URL.
        :param error_label: Виджет для вывода сообщений об ошибках.
        :param tree: Таблица (Treeview) для отображения данных.
        """
        if type(url_entry) != str:
            url = url_entry.get()  # Получаем URL из текстового поля
        else:
            url = url_entry
        if not url:
            error_label.configure(text="Ошибка: URL не может быть пустым")
            return

        error_label.configure(text=f"Получен URL: {url}")

        data = pars_dob(url)  # Парсим данные с помощью функции `pars_dob`
        error_label.configure(text=f"Получены данные: {data}")

        self.create_excel_row(data)  # Добавляем данные в Excel
        error_label.configure(text="Строка успешно добавлена")
        self.load_excel_data(tree)  # Загружаем данные в интерфейс
        

    # Открытие файла Excel
    def open_file(self, tree):
        """
        Открывает диалоговое окно для выбора Excel-файла.
        После выбора файла загружает его содержимое в таблицу интерфейса.
        :param tree: Таблица (Treeview) для отображения данных.
        """
        self._file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if self._file_path:
            self.load_excel_data(tree)

    # Загрузка данных из Excel и отображение их в интерфейсе
    def load_excel_data(self, tree):
        """
        Загружает все строки из Excel-файла и отображает их в таблице интерфейса.
        Очищает старые строки перед загрузкой новых.
        :param tree: Таблица (Treeview) для отображения данных.
        """
        for row in tree.get_children():  # Удаляем старые данные из интерфейса
            tree.delete(row)

        try:
            wb = load_workbook(self._file_path)  # Загружаем книгу Excel
            ws = wb.active
            rows = ws.iter_rows(values_only=True)  # Читаем строки
            for row in rows:
                tree.insert("", "end", values=row)  # Добавляем строки в интерфейс
        except FileNotFoundError:
            pass
